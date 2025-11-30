# utils/excel_processor.py
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from core.config import get_current_vat, get_future_vat


def read_input_excel(path):
    """Простое чтение Excel для импорта из Excel"""
    wb = load_workbook(filename=path, data_only=True)
    sheet = wb.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def write_output_excel_simple(data, path):
    """
    Максимально быстрый и надёжный экспорт в Excel.
    Никаких стилей, цветов, шрифтов — только голые данные.
    Идеально для бухгалтерии и корпоративных сред.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Доп НДС 20 to 22%"

    # Новые 10 колонок — как у тебя в таблице
    headers = [
        "Выполнено",
        "Название договора",
        "№ договора",
        "Сумма договора",
        "Факт на 31.12.2025",
        "Остаток на 2026",
        "Остаток без НДС",
        f"НДС - {int(get_current_vat())}%",
        f"НДС - {int(get_future_vat())}%",
        "Остаток с будущим НДС",
        "Новая стоимость договора",
        "Сумма увеличения по ДС"
    ]

    ws.append(headers)

    for row_data in data:
        row = []
        for header in headers:
            value = row_data.get(header, "")

            # Для итоговой строки оставляем только название и сумму доп. НДС
            if row_data.get("Название договора") == "ИТОГО":
                if header == "Название договора":
                    row.append("ИТОГО")
                elif header == "Сумма увеличения по ДС":
                    row.append(value)
                else:
                    row.append("")
            else:
                row.append(value if value not in (None, "") else "")

        ws.append(row)

    # Только одно улучшение — ширина колонок (это не стиль, это удобно)
    column_widths = [10, 40, 16, 20, 20, 20, 22, 20, 20, 23, 23, 24]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Заморозка шапки — чисто утилитарно
    ws.freeze_panes = "A2"

    try:
        wb.save(path)
        return True
    except Exception as e:
        print(f"Ошибка при сохранении: {e}")
        return False